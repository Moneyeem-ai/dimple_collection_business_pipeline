from typing import Any
import uuid
import base64
import zipfile
from django.db.models.query import QuerySet

from django.shortcuts import redirect
from django.views import generic
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.core.files.base import ContentFile
from django.contrib.auth.mixins import LoginRequiredMixin
from django_pandas.io import read_frame
from django.views.generic import ListView
from django.db.models import Sum
from django.db import transaction

from apps.utils.utils import SideBarSelectedMixin
from apps.product.models import Product, ProductImage, ProductBarcode
from apps.product.forms import (
    ProductForm,
    UploadFileForm,
    BrandSelectionForm,
    ManualFeedingForm,
)
from apps.product.utils import (
    get_or_create_product,
    check_product_is_unique_or_merge,
    clean_extracted_data,
)
from apps.product.tasks import process_image_data

from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
import json
import binascii

from apps.product.models import (
    Product,
    ProductImage,
    ProductBarcode,
    PTFileEntry,
    PTStatus,
    PTFileBatch,
)
from apps.department.models import Department, Brand, Color, Category, SubCategory, Size
from apps.product.serializers import (
    PTFileEntrySerializer,
    PTFileEntryCreateSerializer,
    DepartmentNestedSerializer,
    BrandSerializer,
    SubCategorySerializer,
    CategorySerializer,
    ColorSerializer,
    SizeSerializer,
)
from apps.product.mappers import (
    pt_entry_to_product_mapper,
    pt_entry_to_pt_entry_mapper,
    z_order_upload_to_dict_mapper,
)
from apps.product.tasks import process_image_data


class ProductListView(SideBarSelectedMixin, LoginRequiredMixin, generic.ListView):
    model = Product
    template_name = "pages/product/product_list.html"
    login_url = "user:login"
    context_object_name = "products"
    parent = "product"
    segment = "product_list"
    paginate_by = 6
    ordering = "-created_at"


class ProductBarcodeListView(
    SideBarSelectedMixin, LoginRequiredMixin, generic.ListView
):
    model = PTFileBatch
    template_name = "pages/product/barcode_list.html"
    login_url = "user:login"
    context_object_name = "batch_list"
    parent = "product"
    segment = "barcode_list"
    paginate_by = 6
    ordering = "-created_at"


class ProductEntryView(SideBarSelectedMixin, LoginRequiredMixin, generic.TemplateView):
    template_name = "pages/product/product_entry_photo.html"
    form_class = ProductForm
    parent = "product"
    segment = "product_list"

    def get_context_data(self):
        data = self.request.GET
        context = super().get_context_data()
        context["photo"] = "Cloth" if data.get("page") == "1" else "Tag"
        context["show_fields"] = data.get("page") == "1"
        if data.get("page") == "3":
            id = data.get("id")
            product = Product.objects.get(id=id)
            context["form"] = self.form_class(instance=product)
            context["product"] = product
        return context

    def get(self, request):
        if request.GET.get("page") == "3":
            context = self.get_context_data()
            self.template_name = "pages/product/product_entry.html"
            return self.render_to_response(context=context)
        return super().get(request=request)

    def post(self, request):
        id = request.GET.get("id")
        instance = Product.objects.get(id=id)
        form = self.form_class(request.POST, instance=instance)
        if form.is_valid():
            form.save()
        return redirect("product:product_entry")


class ProductImageView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = request.POST
            image_data = data.get("image").split(",")[1]
            article_data = data.get("article_number")
            color_data = data.get("color")
            print(len(image_data))
            while len(image_data) % 4 != 0:
                image_data += "="
            print(image_data)
            decoded_image_data = base64.b64decode(image_data)
            image_name = str(uuid.uuid4())
            image_file = ContentFile(decoded_image_data, name=f"{image_name}.png")
            product_instance = ProductImage.objects.create(
                product_image=image_file,
                metadata={"article_number": article_data, "color": color_data},
            )
            context = {"product_id": product_instance.id}
            print("product_id", product_instance.id)
            return JsonResponse({"status": "success", "context": context})
        except Exception as e:
            print(e)
            return JsonResponse({"status": "error", "message": str(e)})

    def patch(self, request, *args, **kwargs):
        print("hello")
        product_id = kwargs.get("id")

        # tag_image = request.FILES.get("tag_image")
        try:
            image_data = request.body.decode("utf-8").split(",")[1]
            while len(image_data) % 4 != 0:
                image_data += "="
            decoded_image_data = base64.b64decode(image_data)
            image_name = str(uuid.uuid4())
            image_file = ContentFile(decoded_image_data, name=f"{image_name}.png")
            try:
                product_image_instance = ProductImage.objects.get(id=product_id)
                product_image_instance.tag_image = image_file
                product_image_instance.save()
                tag_image = product_image_instance.tag_image.name.split("/")[1]
                print("tag_image", tag_image)
                process_image_data.delay(tag_image, product_image_instance.id)
                print("product_id_after", "done save product")
                context = {}
            except Exception as e:
                print(e)
                context = {}
            return JsonResponse({"status": "success", "context": context})
        except Exception as e:
            print(e)
            return JsonResponse({"status": "error", "message": str(e)})


class UploadFileView(SideBarSelectedMixin, generic.ListView):
    form_class = UploadFileForm
    model = PTFileBatch
    success_url = "product:barcode_list"
    template_name = "pages/product/barcode_batch_detail.html"
    parent = "product"
    segment = "barcode_list"
    paginate_by = 6

    def get_queryset(self) -> QuerySet[Any]:
        batch_id = self.kwargs.get("batch_id")
        batch = PTFileBatch.objects.get(id=batch_id)
        ptfile_entry_ids = batch.ptfile_entry_ids
        ptfile_entries = PTFileEntry.objects.filter(id__in=ptfile_entry_ids).order_by(
            "product__article_number", "size__size_value"
        )
        if batch.is_file_uploaded:
            queryset = ProductBarcode.objects.filter(pt_entry__in=ptfile_entries)
        else:
            queryset = ptfile_entries
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        upload_form = UploadFileForm()
        batch_id = self.kwargs.get("batch_id")
        batch = PTFileBatch.objects.get(id=batch_id)
        barcode_entries = False
        if batch.is_file_uploaded:
            barcode_entries = True
        context["batch"] = batch
        context["upload_form"] = upload_form
        context["barcode_entries"] = barcode_entries
        return context

    def post(self, request, *args, **kwargs):
        batch_id = kwargs.get("batch_id")
        form = self.form_class(request.POST, request.FILES)
        self.object_list = self.get_queryset()
        try:
            if form.is_valid():
                uploaded_file = request.FILES["file"]
                try:
                    batch = PTFileBatch.objects.get(id=batch_id)
                except Exception as e:
                    print(e)
                    return self.render_to_response(status=status.HTTP_400_BAD_REQUEST)

                ptfile_entry_ids = batch.ptfile_entry_ids
                ptfile_entries = PTFileEntry.objects.filter(id__in=ptfile_entry_ids)
                workbook = load_workbook(uploaded_file, read_only=True)
                sheet = workbook.active
                pt_entry_frequency = {}
                barcode_data_list = []
                errors = []
                for row in sheet.iter_rows(min_row=2):
                    row = [cell.value for cell in row]
                    if row[0] is not None:
                        barcode_row = z_order_upload_to_dict_mapper(row)
                        pt_entry_id = barcode_row.get("pt_entry_id")
                        pt_entry_frequency[pt_entry_id] = (
                            pt_entry_frequency.get(pt_entry_id, 0) + 1
                        )
                        barcode_data_list.append(barcode_row)
                for entry in ptfile_entries:
                    expected_quantity = entry.quantity
                    provided_quantity = pt_entry_frequency.get(entry.id, 0)
                    print(f"{entry.id} -> {expected_quantity} == {provided_quantity}")
                    if expected_quantity != provided_quantity:
                        errors.append(
                            f"Quantity mismatch for PT Entry id: {entry.id}. Expected: {expected_quantity}, Provided: {provided_quantity}"
                        )
                if errors:
                    print(errors)
                    context = self.get_context_data()
                    context["errors"] = errors
                    return self.render_to_response(
                        context, status=status.HTTP_400_BAD_REQUEST
                    )

                product_barcodes = [
                    ProductBarcode(
                        pt_entry_id=barcode_data.get("pt_entry_id"),
                        barcode=barcode_data.get("barcode"),
                    )
                    for barcode_data in barcode_data_list
                ]
                for p in product_barcodes:
                    print(p.__dict__)
                try:
                    from django.db import transaction

                    with transaction.atomic():
                        ProductBarcode.objects.bulk_create(product_barcodes)
                        batch.is_file_uploaded = True
                        batch.save()
                except Exception as e:
                    print("#$^&*()")
                    print(e)
                    error_message = str(e)
                    context = self.get_context_data()
                    context["errors"] = [error_message]
                    print("!!!")
                    print(context["errors"])
                    return self.render_to_response(
                        context, status=status.HTTP_400_BAD_REQUEST
                    )
                return self.render_to_response(self.get_context_data())
            else:
                context_data = self.get_context_data()
                context_data["upload_form"] = form
                return self.render_to_response(context_data)
        except Exception as e:
            print(e)
            return self.render_to_response(form)


class PTFileEntryView(SideBarSelectedMixin, LoginRequiredMixin, generic.TemplateView):
    model = PTFileEntry
    template_name = "pages/product/pt_entry.html"
    parent = "product"
    segment = "ptfile_entry"


class PTFileEntryListView(
    SideBarSelectedMixin, LoginRequiredMixin, generic.TemplateView
):
    model = PTFileEntry
    template_name = "pages/product/pt_list.html"
    parent = "product"
    segment = "batch_list"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        batch_id = kwargs.get("batch_id")
        batch = PTFileBatch.objects.get(id=batch_id)
        ptfile_entry_ids = batch.ptfile_entry_ids
        total_quantity = PTFileEntry.objects.filter(id__in=ptfile_entry_ids).aggregate(
            total_quantity=Sum("quantity")
        )["total_quantity"]
        context["total_quantity"] = total_quantity
        context["batch_id"] = batch_id
        if self.request.user.is_authenticated:
            context["user_type"] = self.request.user.user_type
        else:
            context["user_type"] = None
        return context


class PTFileEntryAPIView(generics.ListAPIView):
    queryset = PTFileEntry.objects.filter(status="ENTRY").order_by(
        "product__article_number", "size__size_value"
    )
    serializer_class = PTFileEntrySerializer

    def list(self, request, *args, **kwargs):
        try:
            print(f"PT Entry API called - Request user: {request.user}")
            queryset = self.filter_queryset(self.get_queryset())
            print(f"Queryset count: {queryset.count()}")

            serializer = self.get_serializer(queryset, many=True)
            departments = DepartmentNestedSerializer(
                Department.objects.all(), many=True
            ).data
            categories = CategorySerializer(Category.objects.all(), many=True).data
            subcategories = SubCategorySerializer(
                SubCategory.objects.all(), many=True
            ).data
            sizes = SizeSerializer(Size.objects.all(), many=True).data
            brands = BrandSerializer(Brand.objects.all(), many=True).data
            colors = ColorSerializer(Color.objects.all(), many=True).data
            data = serializer.data
            print(f"Serialized data count: {len(data)}")

            result = {
                "data": data,
                "departments": departments,
                "sizes": sizes,
                "categories": categories,
                "subcategories": subcategories,
                "brands": brands,
                "colors": colors,
            }
            return Response(result)
        except Exception as e:
            print(f"Error in PTFileEntryAPIView: {e}")
            import traceback

            traceback.print_exc()
            return Response({"error": str(e)}, status=500)


class PTFileEntryListAPIView(generics.ListAPIView):
    serializer_class = PTFileEntrySerializer

    def list(self, request, batch_id, *args, **kwargs):
        print("pt_list_batch_id", batch_id)
        batch = PTFileBatch.objects.get(id=batch_id)
        ptfile_entry_ids = batch.ptfile_entry_ids
        ptfile_entries = PTFileEntry.objects.filter(id__in=ptfile_entry_ids).order_by(
            "product__article_number", "size__size_value"
        )
        serializer = self.get_serializer(ptfile_entries, many=True)
        departments = DepartmentNestedSerializer(
            Department.objects.all(), many=True
        ).data
        brands = BrandSerializer(Brand.objects.all(), many=True).data
        colors = ColorSerializer(Color.objects.all(), many=True).data
        data = serializer.data
        result = {
            "data": data,
            "departments": departments,
            "brands": brands,
            "colors": colors,
        }
        return Response(result)


class PTFileEntryUpdateAPIView(APIView):
    def post(self, request):
        try:
            reques_data = request.data
            pt_file_entries = reques_data.get("data")
            ptstatus = reques_data.get("status")
            batch_id = reques_data.get("id", None)

            with transaction.atomic():
                if batch_id:
                    pt_batch = PTFileBatch.objects.get(id=batch_id)
                    existing_entries = PTFileEntry.objects.filter(
                        status=ptstatus, id__in=pt_batch.ptfile_entry_ids
                    )
                else:
                    existing_entries = PTFileEntry.objects.filter(status=ptstatus)

                existing_ids = [entry.id for entry in existing_entries]
                incoming_ids = [
                    int(entry[0]) if entry[0] else None for entry in pt_file_entries
                ]
                ids_to_delete = list(set(existing_ids) - set(incoming_ids))
                PTFileEntry.objects.filter(id__in=ids_to_delete).delete()
                pt_entry_ids = []

                for pt_entry in pt_file_entries:
                    entry_id = pt_entry[0]
                    if entry_id:
                        pt_file_entry = PTFileEntry.objects.get(id=entry_id)
                        product_data = pt_entry_to_product_mapper(pt_entry)
                        product = pt_file_entry.product
                        for key, value in product_data.items():
                            setattr(product, key, value)
                        product.save()
                        pt_entry_data = pt_entry_to_pt_entry_mapper(
                            pt_entry, batch_id=batch_id
                        )
                        serializer = PTFileEntryCreateSerializer(
                            pt_file_entry, data=pt_entry_data, partial=True
                        )
                    else:
                        product_id = pt_entry[1]
                        product_data = pt_entry_to_product_mapper(
                            pt_entry, without_images=False, without_metadata=False
                        )
                        if product_id is None:
                            product = get_or_create_product(product_data)
                        else:
                            product = Product.objects.get(id=product_id)
                            for key, value in product_data.items():
                                setattr(product, key, value)
                            product.save()
                        instance = PTFileEntry.objects.create(product=product)
                        pt_entry_data = pt_entry_to_pt_entry_mapper(
                            pt_entry, without_product_id=False, batch_id=batch_id
                        )
                        serializer = PTFileEntryCreateSerializer(
                            instance, data=pt_entry_data, partial=True
                        )
                    if serializer.is_valid():
                        serializer.save(status=PTStatus.LIST)
                    else:
                        print("error1", serializer.errors)
                        return Response(
                            serializer.errors,
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        )

                    if pt_entry_id := serializer.data.get("id"):
                        pt_entry_ids.append(pt_entry_id)

                check_product_is_unique_or_merge(pt_entry_ids)

                if ptstatus == PTStatus.ENTRY:
                    if len(pt_entry_ids) > 0 and batch_id is None:
                        try:
                            ptb = PTFileBatch.objects.create(
                                ptfile_entry_ids=pt_entry_ids
                            )
                            if ptb.ptfile_entry_ids != []:
                                pt = PTFileEntry.objects.get(id=ptb.ptfile_entry_ids[0])
                                ptb.batch_id = (
                                    f"{pt.product.brand.brand_name}_{ptb.batch_id}"
                                )
                                ptb.save()
                        except Exception as e:
                            print("error2", e)
                            return Response({"error": str(e)})
                elif batch_id is not None:
                    try:
                        pt_batch = PTFileBatch.objects.get(id=batch_id)
                        pt_batch.ptfile_entry_ids = pt_entry_ids
                        pt_batch.save()
                    except Exception as e:
                        print("error3", e)
                        return Response({"error": str(e)})
            return Response(
                {"message": "Data updated successfully", "success": True},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print("error4", e)
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class BatchListView(SideBarSelectedMixin, ListView):
    model = PTFileBatch
    queryset = PTFileBatch.objects.filter(is_file_uploaded=False)
    template_name = "pages/product/pt_file_batch.html"
    context_object_name = "batch_list"
    parent = "product"
    segment = "batch_list"
    paginate_by = 6
    ordering = "-created_at"


class BaseExportView(View):
    login_url = "users:account_login"

    def get(self, request, *args, **kwargs):
        batch_id = kwargs.get("batch_id")
        try:
            batch = PTFileBatch.objects.get(id=batch_id)
        except PTFileBatch.DoesNotExist:
            # Handle the case where the batch does not exist
            return HttpResponse("Batch not found", status=404)

        # Filter PTFileEntries based on the ptfile_entry_ids field of the PTFileBatch.
        queryset = (
            PTFileEntry.objects.filter(id__in=batch.ptfile_entry_ids)
            .select_related(
                "product",
                "product__department",
                "product__category",
                "product__subcategory",
                "product__brand",
            )
            .order_by("product__article_number", "size__size_value")
        )

        fields_to_export = [
            "product__department__department_name",
            "product__category__category_name",
            "product__subcategory__subcategory_name",
            "product__department__prefix",
            "product__subcategory__prefix",
            "product__category__prefix",
            "product__brand__suffix",
            "product__article_number",
            "id",
            "color_code",
            "color__color_name",
            "size__size_value",
            "product__brand__brand_code",
            "product__category__hsn_code",
            "product__brand__supplier_name",
            "pur_price",
            "mrp",
            "quantity",
            "invoice_number",
            "invoice_date",
        ]
        df = read_frame(queryset, fieldnames=fields_to_export)
        column_mapping = {
            "product__department__department_name": "Department",
            "product__category__category_name": "Category",
            "product__department__prefix": "Department Prefix",
            "product__subcategory__prefix": "Subcategory Prefix",
            "product__category__prefix": "Category Prefix",
            "product__subcategory__subcategory_name": "SubCategory",
            "product__article_number": "ArticleNo",
            "id": "Description",
            "color_code": "ExtDescription",
            "color__color_name": "Color",
            "size__size_value": "Size",
            "product__brand__suffix": "Brand Suffix",
            "product__brand__brand_code": "Brand",
            "product__category__hsn_code": "HSNCode",
            "product__brand__supplier_name": "Supplier",
            "pur_price": "PurPrice",
            "mrp": "ItemMrp",
            "quantity": "Quantity",
            "invoice_number": "InvoiceNo",
            "invoice_date": "InvoiceDt",
        }
        df = df.rename(columns=column_mapping)
        df["InvoiceDt"] = pd.to_datetime(df["InvoiceDt"])
        df["InvoiceDt"] = df["InvoiceDt"].dt.strftime("%d-%m-%Y")
        df["Brand Suffix"].fillna("", inplace=True)
        df["Department Prefix"].fillna("", inplace=True)
        df["Category Prefix"].fillna("", inplace=True)
        df["Subcategory Prefix"].fillna("", inplace=True)
        df["ArticleNo"] = (
            df["Department Prefix"].astype(str)
            + df["Category Prefix"].astype(str)
            + df["Subcategory Prefix"].astype(str)
            + df["ArticleNo"].astype(str)
            + df["Brand Suffix"].astype(str)
        )
        df.drop(
            columns=[
                "Department Prefix",
                "Category Prefix",
                "Subcategory Prefix",
                "Brand Suffix",
            ],
            inplace=True,
        )

        df.insert(4, "CodingType", 3)
        df.insert(5, "UOMName", "pcs")
        df.insert(10, "Style", None)
        df.insert(14, "ItemCode", None)
        df.insert(15, "ItemId", None)
        df.insert(18, "ItemWsp", None)
        df.insert(22, "PORowId", None)
        df.insert(23, "PurOrderId", None)

        csv_file_path = "data/ptfiles_export.csv"
        df.to_csv(csv_file_path, index=False)

        self.update_batch(batch)

        queryset = PTFileEntry.objects.filter(id__in=batch.ptfile_entry_ids)
        ptfile_entry = queryset.first()
        product = ptfile_entry.product
        brand = product.brand
        brand_name = brand.brand_name
        file_name = (
            f'{brand_name}_{pd.Timestamp.now().strftime("%Y-%m-%d_%H-%M-%S")}.csv'
        )

        with open(csv_file_path, "rb") as csv_file:
            response = HttpResponse(
                csv_file.read(),
                content_type="text/csv",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'

        return response

    def update_batch(self, batch):
        raise NotImplementedError("Subclasses must implement this method")


class ExportPTFilesView(BaseExportView):
    def update_batch(self, batch):
        batch.is_exported = True
        batch.save()


class MarkPOExportedView(BaseExportView):
    def update_batch(self, batch):
        batch.is_exported_for_po = True
        batch.save()


class ExportImagesAPIView(View):
    def get(self, request, batch_id):
        batch = PTFileBatch.objects.get(id=batch_id)
        ptfile_entry_ids = batch.ptfile_entry_ids
        ptfile_entries = PTFileEntry.objects.filter(id__in=ptfile_entry_ids)

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zip_file:
            for ptfile_entry in ptfile_entries:
                product = ptfile_entry.product
                product_images = product.product_images
                if product_images.product_image:
                    image_path = str(product_images.product_image.path)
                    image_name = ""
                    if product.department.prefix is not None:
                        image_name += product.department.prefix
                    if product.category.prefix is not None:
                        image_name += product.category.prefix
                    if product.subcategory.prefix is not None:
                        image_name += product.subcategory.prefix
                    image_name += f"{product.article_number}"
                    if product.brand.suffix is not None:
                        image_name += product.brand.suffix
                    zip_file.write(image_path, arcname=f"{image_name}.jpg")

        batch.is_image_exported = True
        batch.save()

        # Build the response with the zip file contents
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = (
            f"attachment; filename=product_images_batch_{batch_id}.zip"
        )
        return response


class ManualFeedingBrandSelectionView(
    SideBarSelectedMixin, LoginRequiredMixin, generic.FormView
):
    template_name = "pages/product/manual_feeding_brand_selection.html"
    form_class = BrandSelectionForm
    parent = "product"
    segment = "manual_feeding"

    def form_valid(self, form):
        brand_id = form.cleaned_data["brand"].id
        return redirect("product:manual_feeding_form", brand_id=brand_id)


class ManualFeedingFormView(SideBarSelectedMixin, LoginRequiredMixin, generic.FormView):
    template_name = "pages/product/manual_feeding_form.html"
    form_class = ManualFeedingForm
    parent = "product"
    segment = "manual_feeding"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        brand_id = self.kwargs.get("brand_id")
        try:
            brand = Brand.objects.get(id=brand_id)
            context["brand"] = brand
        except Brand.DoesNotExist:
            context["brand"] = None
        return context

    def post(self, request, *args, **kwargs):
        """Override post to add debugging"""
        print("=== MANUAL FEEDING POST REQUEST ===")
        print(f"Request method: {request.method}")
        print(f"Request POST data: {request.POST}")
        print(f"Files: {request.FILES}")
        print(f"Session data: {dict(request.session.items())}")
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        print("=== FORM_VALID CALLED ===")
        brand_id = self.kwargs.get("brand_id")
        try:
            brand = Brand.objects.get(id=brand_id)
        except Brand.DoesNotExist:
            return redirect("product:manual_feeding_brand_selection")

        department = form.cleaned_data["department"]
        article = form.cleaned_data["article"]
        color = form.cleaned_data["color"]
        number_of_rows = form.cleaned_data["number_of_rows"]
        photo_data = form.cleaned_data.get("photo_data")

        print("number_of_rows", number_of_rows)
        print("article", article)
        print("color", color)
        print("department", department)
        print("photo_data length:", len(photo_data) if photo_data else "None")

        try:
            format, imgstr = photo_data.split(";base64,")
            ext = format.split("/")[-1]
            image_file = ContentFile(
                base64.b64decode(imgstr), name=f"{article}_{uuid.uuid4()}.{ext}"
            )
        except Exception as e:
            form.add_error(None, f"Error processing image: {str(e)}")
            return self.form_invalid(form)

        try:
            product_image = ProductImage.objects.create(
                product_image=image_file,
                metadata={
                    "manual_feeding": True,
                    "article_number": article.upper(),
                    "color": color.color_name,
                },
            )
            print(f"Created ProductImage: {product_image.id}")
            extracted_data = {
                "department_id": department.id,
                "brand_id": brand.id,
                "color_id": color.id,
                "article_number": article.upper(),
                "mrp": 0,  # Default MRP for manual feeding
            }
            print("Extracted data:", extracted_data)

            valid_data = clean_extracted_data(extracted_data, product_image.id)
            print("Valid data:", valid_data)

            product = get_or_create_product(valid_data)
            print(f"Product created/retrieved: {product.id} - {product.article_number}")

            pt_entry_ids = []
            with transaction.atomic():
                for i in range(number_of_rows):
                    pt_entry = PTFileEntry.objects.create(product=product)
                    pt_entry_ids.append(pt_entry.id)
                    print(f"Created PT Entry: {pt_entry.id}")
            return redirect(f"/product/manual-feeding/form/{brand_id}/")

        except Exception as e:
            import traceback

            print("Error in manual feeding:", str(e))
            print(traceback.format_exc())
            form.add_error(None, f"Error creating product: {str(e)}")
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Handle form validation errors"""
        print("=== FORM_INVALID CALLED ===")
        print(f"Form errors: {form.errors}")
        print(f"Form non-field errors: {form.non_field_errors()}")
        for field_name, field in form.fields.items():
            field_value = form.data.get(field_name, "NOT PROVIDED")
            print(f"Field {field_name}: {field_value}")
        return super().form_invalid(form)


class ArticleAutocompleteAPIView(APIView):
    """
    API endpoint for article number autocomplete suggestions.
    Returns article numbers that match the search term for a specific brand.
    """

    def get(self, request, brand_id):
        search_term = request.GET.get("q", "").strip()

        if len(search_term) < 2:  # Only search if user typed at least 2 characters
            return Response({"suggestions": []})

        try:
            # Get products for the specific brand that match the search term
            products = (
                Product.objects.filter(
                    brand_id=brand_id, article_number__icontains=search_term
                )
                .values_list("article_number", flat=True)
                .distinct()[:10]
            )  # Limit to 10 suggestions

            suggestions = list(products)

            return Response({"suggestions": suggestions, "count": len(suggestions)})

        except Exception as e:
            return Response(
                {"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
